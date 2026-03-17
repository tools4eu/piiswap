"""Excel (.xlsx) adapter using openpyxl."""

from pathlib import Path
from typing import Callable, List, Optional

from piiswap.adapters.base import FileAdapter, register_adapter


@register_adapter
class XlsxAdapter(FileAdapter):
    """Read and write Excel (.xlsx) files.

    Reads all text content from all sheets for PII detection.
    Write creates an anonymized/de-anonymized copy preserving formatting.
    """

    supported_extensions = (".xlsx", ".xls")

    def read(self, path: Path) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"--- Sheet: {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                line = "\t".join(cells)
                if line.strip():
                    parts.append(line)

        wb.close()
        return "\n".join(parts)

    def write(self, path: Path, content: str) -> None:
        """Write as plain text (fallback when no source workbook available)."""
        path.parent.mkdir(parents=True, exist_ok=True)
        txt_path = path.with_suffix(".txt")
        txt_path.write_text(content, encoding="utf-8")

    def supports_columns(self) -> bool:
        return True

    def anonymize_preserving_format(
        self,
        input_path: Path,
        output_path: Path,
        replace_fn: Callable[[str], str],
        pii_columns: Optional[List[str]] = None,
        keep_columns: Optional[List[str]] = None,
    ) -> None:
        """Create an anonymized/de-anonymized copy preserving Excel formatting.

        When ``pii_columns`` or ``keep_columns`` are supplied, column filtering
        is applied per sheet using the values in row 1 as header names.

        Args:
            input_path:  Source .xlsx file.
            output_path: Destination .xlsx file.
            replace_fn:  Callable that takes a cell string and returns the
                         anonymized/de-anonymized version.
            pii_columns: Only process cells in these columns (matched by
                         header name in row 1 of each sheet).
            keep_columns: Process all columns EXCEPT these.
        """
        from openpyxl import load_workbook

        wb = load_workbook(str(input_path))

        for ws in wb.worksheets:
            # Resolve column filtering for this sheet
            target_col_indices = _resolve_xlsx_target_columns(
                ws, pii_columns, keep_columns
            )

            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                for cell in row:
                    # Skip the header row itself from being replaced
                    if row_idx == 1 and (pii_columns or keep_columns):
                        continue
                    # Skip if column filtering is active and this col is excluded
                    if target_col_indices is not None and cell.column not in target_col_indices:
                        continue
                    if cell.value and isinstance(cell.value, str):
                        new_val = replace_fn(cell.value)
                        if new_val != cell.value:
                            cell.value = new_val

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        wb.close()


def _resolve_xlsx_target_columns(
    ws,
    pii_columns: Optional[List[str]],
    keep_columns: Optional[List[str]],
) -> Optional[set]:
    """Return a set of 1-based column indices to process, or None for 'all'.

    Returns None when no column filtering should be applied (process everything).
    """
    if not pii_columns and not keep_columns:
        return None  # No filtering — process all cells

    # Read header names from row 1 (case-insensitive matching)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), [])
    col_lower_to_index = {}
    for cell in header_row:
        if cell.value is not None:
            col_lower_to_index[str(cell.value).lower()] = cell.column

    if pii_columns:
        return {col_lower_to_index[c.lower()] for c in pii_columns if c.lower() in col_lower_to_index}

    # keep_columns: process everything except those columns
    keep_lower = {c.lower() for c in keep_columns}
    return {idx for idx in col_lower_to_index.values()
            if list(col_lower_to_index.keys())[list(col_lower_to_index.values()).index(idx)] not in keep_lower}
